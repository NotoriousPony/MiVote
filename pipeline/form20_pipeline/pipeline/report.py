"""Validation report workbook + readable CSV exports."""
import os, json, csv, sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def run(db_path, cfg):
    out = cfg.OUTPUT_DIR
    report = json.load(open(os.path.join(out, 'batch_report.json')))
    con = sqlite3.connect(db_path)

    with open(os.path.join(out, 'candidates.csv'), 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['assembly', 'candidate_index', 'candidate_name', 'party', 'total_votes_evm', 'rank', 'is_top4'])
        w.writerows(con.execute('SELECT ac_name, cand_idx, name, party, total_votes_evm, rank, is_top4 FROM candidate ORDER BY ac_name, rank'))
    with open(os.path.join(out, 'village_results_readable.csv'), 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['assembly', 'village', 'candidate', 'party', 'votes', 'rank_in_ac', 'is_top4'])
        w.writerows(con.execute('''SELECT r.ac_name, r.village, c.name, c.party, r.votes, c.rank, c.is_top4
            FROM result r JOIN candidate c ON r.ac_name=c.ac_name AND r.cand_idx=c.cand_idx
            ORDER BY r.ac_name, r.village, c.rank'''))

    wb = openpyxl.Workbook()
    F = 'Arial'
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    warn_fill = PatternFill('solid', fgColor='FCE4EC')
    ok_fill = PatternFill('solid', fgColor='E8F5E9')
    thin = Border(*[Side(style='thin', color='BFBFBF')] * 4)

    def header(ws, cols):
        ws.append(cols)
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(name=F, bold=True, color='FFFFFF', size=10)
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.freeze_panes = 'A2'

    ws = wb.active
    ws.title = 'AC Status'
    header(ws, ['#', 'Constituency', 'Booths', 'Candidates', 'Villages', 'Rows failing checksum',
                'Totals match Form 20', 'Unmapped booths', 'Status / action needed'])
    for i, e in enumerate(sorted(report, key=lambda x: x['ac']), start=1):
        if e['status'] != 'OK':
            note = cfg.MANUAL_ENTRIES.get(e['ac'], {}).get('status', e['note'])
            ws.append([i, e['ac'], None, None, None, None, None, None, note])
            bad = e['ac'] not in cfg.MANUAL_ENTRIES
        else:
            bad = bool(e['unmapped_booths'])
            action = 'OK' if not bad else 'Add booths to mapping (%d missing: %s...)' % (
                e['unmapped_booths'], ', '.join(e['unmapped_list'][:6]))
            ws.append([i, e['ac'], e['booths'], e['ncand'], e['villages'], e['internal_bad'],
                       'YES' if e['evm_match'] else 'NO', e['unmapped_booths'], action])
        rn = ws.max_row
        for c in range(1, 10):
            cell = ws.cell(row=rn, column=c)
            cell.font = Font(name=F, size=10)
            cell.border = thin
            cell.fill = warn_fill if bad else ok_fill
    ws.column_dimensions['B'].width = 20
    for col in 'CDEFGH':
        ws.column_dimensions[col].width = 12
    ws.column_dimensions['I'].width = 55

    ws2 = wb.create_sheet('Candidates')
    header(ws2, ['Constituency', 'Rank', 'Candidate', 'Party', 'Total EVM votes', 'Top 4?'])
    for ac, rank, name, party, votes, top4 in con.execute(
            'SELECT ac_name, rank, name, party, total_votes_evm, is_top4 FROM candidate ORDER BY ac_name, rank'):
        ws2.append([ac, rank, name, party, votes, 'YES' if top4 else ''])
        rn = ws2.max_row
        for c in range(1, 7):
            cell = ws2.cell(row=rn, column=c)
            cell.font = Font(name=F, size=10)
            cell.border = thin
            if c == 5:
                cell.number_format = '#,##0'
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['C'].width = 32
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14

    ws3 = wb.create_sheet('Notes')
    for i, n in enumerate(cfg.REPORT_NOTES, start=1):
        ws3.cell(row=i, column=1, value=n).font = Font(name=F, size=10, bold=(i == 1))
    ws3.column_dimensions['A'].width = 120

    wb.save(os.path.join(out, 'validation_report.xlsx'))
    con.close()
    print('report + CSVs written to', out)
