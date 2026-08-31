"""Join booth votes with booth->village mapping, aggregate, build SQLite DB."""
import os, re, json, csv, sqlite3, difflib, tempfile, shutil
import openpyxl
from collections import defaultdict
from .parse import parse_form20
from .names import extract_candidate_names


def norm(s):
    s = s.lower().replace('(sc)', '').replace('(st)', '').replace('.pdf', '')
    return re.sub(r'[^a-z]', '', s)


def match_pdfs_to_sheets(pdf_dir, sheets, aliases):
    pdfs = sorted(f for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf'))
    sheet_by_norm = {norm(s): s for s in sheets}
    pairs, unmatched, used = [], [], set()
    for f in pdfs:
        n = norm(f)
        m = aliases.get(n) or sheet_by_norm.get(n)
        if not m:
            cand = difflib.get_close_matches(
                n, [k for k in sheet_by_norm if sheet_by_norm[k] not in used], n=1, cutoff=0.75)
            m = sheet_by_norm[cand[0]] if cand else None
        if m:
            pairs.append((f, m)); used.add(m)
        else:
            unmatched.append(f)
    return pairs, unmatched, [s for s in sheets if s not in used]


def load_mapping(wb, sheet, cfg):
    b2v = {}
    for a, b, vname in cfg.RANGE_FIXES.get(sheet, []):
        for k in range(a, b + 1):
            b2v[k] = vname
    if sheet not in cfg.SHEET_SKIP:
        for r in wb[sheet].iter_rows(min_row=2, values_only=True):
            if r[0] is not None and r[1] is not None:
                key = r[0]
                try:
                    key = int(key)
                except (ValueError, TypeError):
                    key = str(key).strip().upper()
                b2v[key] = str(r[1]).strip()
    return b2v


def run(cfg):
    out = cfg.OUTPUT_DIR
    os.makedirs(out, exist_ok=True)
    cache_dir = os.path.join(out, 'parse_cache')
    os.makedirs(cache_dir, exist_ok=True)
    ocr_dir = os.path.join(out, 'ocr_cache')

    wb = openpyxl.load_workbook(cfg.MAPPING_XLSX, read_only=True)
    pairs, unmatched_pdfs, unused_sheets = match_pdfs_to_sheets(cfg.PDF_DIR, wb.sheetnames, cfg.AC_ALIASES)
    print('matched: %d | unmatched pdfs: %s | unused sheets: %s' % (len(pairs), unmatched_pdfs, unused_sheets))

    report, agg_out, cand_names = [], {}, {}
    for f, sheet in pairs:
        cf = os.path.join(cache_dir, f.rsplit('.', 1)[0] + '.json')
        if os.path.exists(cf):
            res = json.load(open(cf))
        else:
            res = parse_form20(os.path.join(cfg.PDF_DIR, f), ocr_dir)
            json.dump(res, open(cf, 'w'))
        if 'error' in res:
            report.append({'ac': sheet, 'pdf': f, 'status': 'ERROR', 'note': res['error']})
            continue
        if not res['evm_match'] and res.get('evm'):  # recompute with tolerant check (old caches)
            nc_ = res['ncand']
            bs = [sum(r[2 + i] for r in res['rows']) for i in range(nc_)]
            vs = sum(r[2 + nc_] for r in res['rows'])
            expected = ''.join(str(x) for x in bs + [vs])
            actual = ''.join(str(x) for x in res['evm'][:nc_ + 1])
            res['evm_match'] = actual.startswith(expected)
        b2v = load_mapping(wb, sheet, cfg)

        def lookup(ps):
            if ps in b2v:
                return b2v[ps]
            if isinstance(ps, str):
                m = re.match(r'\d+', ps)
                if m and int(m.group()) in b2v:
                    return b2v[int(m.group())]
            return None

        rows, nc = res['rows'], res['ncand']
        unmapped = sorted({str(r[1]) for r in rows if lookup(r[1]) is None})
        agg = defaultdict(lambda: [0] * (nc + 3))
        for r in rows:
            v = lookup(r[1]) or 'UNMAPPED'
            a = agg[v]
            for i in range(nc):
                a[i] += r[2 + i]
            a[nc] += r[2 + nc]; a[nc + 1] += r[4 + nc]; a[nc + 2] += 1
        agg_out[sheet] = (nc, dict(agg))
        cand_names[sheet] = extract_candidate_names(os.path.join(cfg.PDF_DIR, f), nc)
        report.append({'ac': sheet, 'pdf': f, 'status': 'OK', 'ocred': res['ocred'],
                       'booths': len(rows), 'ncand': nc, 'internal_bad': res['internal_bad'],
                       'evm_match': res['evm_match'], 'unmapped_booths': len(unmapped),
                       'unmapped_list': unmapped[:20], 'villages': len(agg)})

    ok = [r for r in report if r['status'] == 'OK']
    clean = [r for r in ok if r['internal_bad'] == 0 and r['evm_match'] and r['unmapped_booths'] == 0]
    print('\n=== %d/%d parsed | %d fully clean ===' % (len(ok), len(pairs), len(clean)))
    for r in report:
        bad = not (r['status'] == 'OK' and r['internal_bad'] == 0 and r['evm_match'] and r['unmapped_booths'] == 0)
        if r['status'] == 'OK':
            print('%-24s booths=%4d cand=%2d villages=%3d badrows=%d evm=%s unmapped=%d%s' % (
                r['ac'], r['booths'], r['ncand'], r['villages'], r['internal_bad'],
                r['evm_match'], r['unmapped_booths'], '  <<< CHECK' if bad else ''))
        else:
            print('%-24s ERROR: %s  <<< CHECK' % (r['ac'], r['note']))
    json.dump(report, open(os.path.join(out, 'batch_report.json'), 'w'), indent=1)

    # SQLite (write to temp dir first: sqlite on network mounts can fail)
    tmp_db = os.path.join(tempfile.gettempdir(), 'form20_results.db')
    if os.path.exists(tmp_db):
        os.remove(tmp_db)
    con = sqlite3.connect(tmp_db)
    con.executescript('''
    CREATE TABLE assembly (ac_name TEXT PRIMARY KEY, booths INTEGER, villages INTEGER,
      n_candidates INTEGER, data_status TEXT);
    CREATE TABLE candidate (ac_name TEXT, cand_idx INTEGER, name TEXT, party TEXT,
      total_votes_evm INTEGER, rank INTEGER, is_top4 INTEGER, PRIMARY KEY (ac_name, cand_idx));
    CREATE TABLE village (ac_name TEXT, village TEXT, booths INTEGER, valid_votes INTEGER,
      nota INTEGER, PRIMARY KEY (ac_name, village));
    CREATE TABLE result (ac_name TEXT, village TEXT, cand_idx INTEGER, votes INTEGER,
      PRIMARY KEY (ac_name, village, cand_idx));''')

    for e in report:
        if e['status'] != 'OK':
            me = cfg.MANUAL_ENTRIES.get(e['ac'])
            if me:
                con.execute('INSERT INTO assembly VALUES (?,?,?,?,?)',
                            (e['ac'], None, None, len(me['candidates']), me['status']))
                for i, (n, p, v) in enumerate(me['candidates']):
                    con.execute('INSERT INTO candidate VALUES (?,?,?,?,?,?,?)',
                                (e['ac'], i, n, p, v, i + 1, 1 if i < 4 else 0))
            else:
                con.execute('INSERT INTO assembly VALUES (?,?,?,?,?)',
                            (e['ac'], None, None, None, 'MISSING: ' + e['note']))
            continue
        status = 'OK' if not e['unmapped_booths'] else 'PARTIAL: %d booths unmapped' % e['unmapped_booths']
        con.execute('INSERT INTO assembly VALUES (?,?,?,?,?)',
                    (e['ac'], e['booths'], e['villages'], e['ncand'], status))
        nc, agg = agg_out[e['ac']]
        totals = [sum(a[i] for a in agg.values()) for i in range(nc)]
        order = sorted(range(nc), key=lambda i: -totals[i])
        rank = {ci: r + 1 for r, ci in enumerate(order)}
        names = cand_names[e['ac']]
        for ci in range(nc):
            nm = names[ci].title() if names[ci] else 'Candidate %d' % (ci + 1)
            con.execute('INSERT INTO candidate VALUES (?,?,?,?,?,?,?)',
                        (e['ac'], ci, nm, None, totals[ci], rank[ci], 1 if rank[ci] <= 4 else 0))
        for v, a in agg.items():
            con.execute('INSERT INTO village VALUES (?,?,?,?,?)', (e['ac'], v, a[nc + 2], a[nc], a[nc + 1]))
            for ci in range(nc):
                con.execute('INSERT INTO result VALUES (?,?,?,?)', (e['ac'], v, ci, a[ci]))
    con.commit()

    n_v = con.execute('SELECT COUNT(*) FROM village').fetchone()[0]
    n_r = con.execute('SELECT COUNT(*) FROM result').fetchone()[0]
    tv = con.execute('SELECT SUM(votes) FROM result').fetchone()[0]
    print('\nDB: %d villages | %d result rows | %s total EVM votes' % (n_v, n_r, format(tv or 0, ',')))
    con.close()
    shutil.copy(tmp_db, os.path.join(out, 'results.db'))
    return tmp_db, report
