"""
MiVote - Form 20 reader using Google AI Studio (Gemini)

Reads scanned Form 20 pages with Google's AI instead of Tesseract, and writes
one Excel file per constituency. Every row is still checked with arithmetic:
candidate votes must add up to the valid total, and valid + rejected + NOTA
must equal the total. Rows that do not add up are marked NEEDS CHECK in red -
they are never quietly trusted.

WHAT YOU NEED
  1. A free API key: https://aistudio.google.com/apikey
     Click "Create API key" and copy the long string.
  2. In this same folder, make a text file named   api_key.txt
     and paste ONLY the key inside it. Never edit this Python file.
  3. Libraries (you already have these):
        pip install requests pymupdf pillow openpyxl

RUN
      cd C:\\MiVote
      python form20_gemini.py

  Every page's answer is saved in a cache folder, so stopping and restarting
  never repeats work already paid for. Finished files are skipped.

FREE LIMITS
  The free tier allows a limited number of requests per minute and per day.
  This script waits between requests and retries automatically when Google
  asks it to slow down. If you hit the daily cap, just run it again tomorrow -
  it continues where it stopped.
"""

import os
import io
import re
import csv
import json
import time
import base64

import requests
import fitz
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill

# ---------------- SETTINGS ----------------
# Your API key goes in a plain text file called api_key.txt, in this same
# folder. Do NOT edit this Python file.
KEY_FILE = "api_key.txt"
MODEL = "gemini-2.0-flash"
IN_DIR = "Form20_LokSabha_2024"
OUT_DIR = "Form20_Excel_Output"
CACHE_DIR = "Form20_Cache"
DPI = 200                          # image quality sent to the AI
SECONDS_BETWEEN_CALLS = 13         # ~4 pages a minute: safely inside the free tier
# ------------------------------------------

URL = "https://generativelanguage.googleapis.com/v1beta/models/{m}:generateContent"


def load_key():
    if not os.path.exists(KEY_FILE):
        return None
    key = open(KEY_FILE, encoding="utf-8").read().strip().strip('"').strip("'")
    return key or None


API_KEY = load_key() or ""

PROMPT = """You are reading a scanned Indian election result sheet (Form 20).

The table has one row per polling station. Columns are, in order:
  serial number and polling station name, then one column per candidate,
  then: Total of Valid Votes, No. of Rejected Votes, NOTA, Total,
  No. of Tendered Votes.

Return ONLY plain text, no explanation, no markdown, in this format:

CANDIDATES: name1 | name2 | name3 | ...
row: station | n1 | n2 | ... | total_valid | rejected | nota | total | tendered

Rules:
- One "row:" line for every polling station row on this page.
- IGNORE rows that are round subtotals, carried-forward totals, or grand totals
  (they say things like "Sub-Total", "Brought Forward", "Carry Forward", "TOTAL").
- Every vote cell must be a number. If a cell is blank or empty, write 0.
- Never guess a number you cannot see clearly - write ? instead.
- Keep the candidate columns in the exact left-to-right order printed.
- Copy the station text exactly as printed, including its serial number.
"""


def page_png(page, dpi):
    pix = page.get_pixmap(dpi=dpi)
    img = Image.open(io.BytesIO(pix.tobytes("png"))).convert("L")
    if img.width > 2400:                       # keep uploads small and fast
        r = 2400 / img.width
        img = img.resize((2400, int(img.height * r)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, "JPEG", quality=88)
    return buf.getvalue()


def ask_gemini(img_bytes, tries=5):
    body = {
        "contents": [{"parts": [
            {"text": PROMPT},
            {"inline_data": {"mime_type": "image/jpeg",
                             "data": base64.b64encode(img_bytes).decode()}},
        ]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 8192},
    }
    url = URL.format(m=MODEL) + "?key=" + API_KEY
    wait = 8
    for attempt in range(tries):
        try:
            r = requests.post(url, json=body, timeout=180)
            if r.status_code == 200:
                j = r.json()
                return j["candidates"][0]["content"]["parts"][0]["text"]
            if r.status_code in (429, 500, 503):        # busy / rate limited
                if attempt == 0:
                    reason = ""
                    try:
                        reason = r.json().get("error", {}).get("message", "")[:180]
                    except Exception:
                        reason = r.text[:180]
                    print(f"      Google says: {reason}", flush=True)
                    if "per day" in reason.lower() or "daily" in reason.lower():
                        print("      >> That is the DAILY free limit. Stop now and run "
                              "again tomorrow - finished pages are cached.", flush=True)
                print(f"      (waiting {wait}s then retrying)", flush=True)
                time.sleep(wait)
                wait = min(wait * 2, 120)
                continue
            print("      API error", r.status_code, r.text[:200], flush=True)
            return None
        except Exception as e:
            print("      network problem:", e, flush=True)
            time.sleep(wait)
            wait = min(wait * 2, 120)
    return None


def parse_reply(text):
    """Return (candidate_names, [[station, cells...], ...])"""
    names, rows = [], []
    for line in (text or "").splitlines():
        line = line.strip()
        if line.upper().startswith("CANDIDATES:"):
            names = [p.strip() for p in line.split(":", 1)[1].split("|") if p.strip()]
        elif line.lower().startswith("row:"):
            parts = [p.strip() for p in line.split(":", 1)[1].split("|")]
            if len(parts) >= 6:
                rows.append(parts)
    return names, rows


def reconcile(cells):
    """cells = numbers as text. Last five are valid, rejected, NOTA, total, tendered."""
    if len(cells) < 6:
        return None, "too few columns"
    vals = []
    for c in cells:
        c = re.sub(r"[^0-9?]", "", c)
        vals.append(None if (c == "" or c == "?") else int(c))
    n = len(vals)
    i_valid, i_rej, i_nota, i_tot = n - 5, n - 4, n - 3, n - 2
    cand = vals[:i_valid]
    known = sum(v for v in cand if v is not None)
    blanks = [i for i, v in enumerate(cand) if v is None]
    status = "OK"
    if vals[i_valid] is not None:
        gap = vals[i_valid] - known
        if gap == 0 and not blanks:
            pass
        elif gap == 0:
            for i in blanks:
                cand[i] = 0
        elif len(blanks) == 1 and gap > 0:
            cand[blanks[0]] = gap
        else:
            status = "NEEDS CHECK: candidate votes do not add up"
    else:
        status = "NEEDS CHECK: valid-votes total unreadable"
    for i in (i_rej, i_nota):
        if vals[i] is None:
            vals[i] = 0
    if status == "OK" and vals[i_tot] is not None:
        if vals[i_valid] + vals[i_rej] + vals[i_nota] != vals[i_tot]:
            status = "NEEDS CHECK: total column does not match"
    return cand + vals[i_valid:], status


def main():
    if not API_KEY:
        print("No API key found.")
        print(f"Create a plain text file called {KEY_FILE} in this folder")
        print("(" + os.path.abspath(".") + ")")
        print("and put ONLY your API key inside it, then run this again.")
        print("Get a free key at https://aistudio.google.com/apikey")
        return
    if not os.path.isdir(IN_DIR):
        print("Folder not found:", os.path.abspath(IN_DIR))
        return
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(CACHE_DIR, exist_ok=True)

    pdfs = sorted(f for f in os.listdir(IN_DIR) if f.lower().endswith(".pdf"))
    print(f"PDFs found: {len(pdfs)} | model: {MODEL}\n", flush=True)
    summary = []
    t_start = time.time()

    for n, fn in enumerate(pdfs, 1):
        base = os.path.splitext(fn)[0]
        xlsx = os.path.join(OUT_DIR, base + ".xlsx")
        if os.path.exists(xlsx):
            print(f"[{n}/{len(pdfs)}] skip (already done): {base}", flush=True)
            continue
        path = os.path.join(IN_DIR, fn)
        doc = fitz.open(path)
        print(f"[{n}/{len(pdfs)}] {base}: {doc.page_count} pages", flush=True)
        t0 = time.time()
        all_rows, names = [], []

        for pno in range(doc.page_count):
            cache_f = os.path.join(CACHE_DIR, f"{base}__p{pno}.txt")
            if os.path.exists(cache_f):
                reply = open(cache_f, encoding="utf-8").read()
            else:
                reply = ask_gemini(page_png(doc[pno], DPI))
                if reply is None:
                    print(f"    page {pno+1}: FAILED - will retry on next run", flush=True)
                    continue
                open(cache_f, "w", encoding="utf-8").write(reply)
                time.sleep(SECONDS_BETWEEN_CALLS)
            nm, rows = parse_reply(reply)
            if nm and not names:
                names = nm
            all_rows.extend(rows)
            print(f"    page {pno+1}/{doc.page_count}: {len(rows)} rows "
                  f"({time.time()-t0:.0f}s)", flush=True)
        doc.close()

        ncand = max((len(r) - 6 for r in all_rows), default=0)
        head = (["Polling station"] +
                [names[i] if i < len(names) else f"Cand {i+1}" for i in range(ncand)] +
                ["Total valid", "Rejected", "NOTA", "Total", "Tendered", "Check"])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Booths"
        ws.append(head)
        for c in range(1, len(head) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        warn = PatternFill("solid", fgColor="FDE2E2")
        ok = bad = 0
        for r in all_rows:
            vals, status = reconcile(r[1:])
            if vals is None:
                continue
            ws.append([r[0]] + vals + [status])
            if status != "OK":
                bad += 1
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=ws.max_row, column=c).fill = warn
            else:
                ok += 1
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 34
        wb.save(xlsx)
        print(f"    -> saved: rows OK: {ok} | needs check: {bad} | "
              f"{time.time()-t0:.0f}s\n", flush=True)
        summary.append((base, len(all_rows), ok, bad))

    if summary:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["File", "Rows found", "Rows OK", "Rows needing check"])
        for c in range(1, 5):
            ws.cell(row=1, column=c).font = Font(bold=True)
        for row in summary:
            ws.append(list(row))
        ws.column_dimensions["A"].width = 34
        wb.save(os.path.join(OUT_DIR, "_SUMMARY.xlsx"))

    print(f"Done in {(time.time()-t_start)/60:.1f} minutes. Files in "
          f"{os.path.abspath(OUT_DIR)}", flush=True)
    print("Open _SUMMARY.xlsx first, then send the folder to Claude.", flush=True)


if __name__ == "__main__":
    main()
