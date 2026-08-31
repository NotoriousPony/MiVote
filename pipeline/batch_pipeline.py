"""Batch pipeline: 90 Form 20 PDFs -> village-wise master database + validation report."""
import fitz, re, os, json, difflib, csv
import openpyxl
from collections import defaultdict

OUT = '/sessions/inspiring-charming-cerf/mnt/outputs/'
PDF_DIR = OUT + 'form20/'
XLSX = '/sessions/inspiring-charming-cerf/mnt/uploads/Haryana_Assembly_Normalized.xlsx'

def norm(s):
    s = s.lower().replace('(sc)', '').replace('.pdf', '')
    return re.sub(r'[^a-z]', '', s)

# --- match PDFs to sheets ---
wb_map = openpyxl.load_workbook(XLSX, read_only=True)
sheets = wb_map.sheetnames
pdfs = sorted(f for f in os.listdir(PDF_DIR) if f.lower().endswith('.pdf'))
sheet_by_norm = {norm(s): s for s in sheets}
ALIASES = {'gurgaon': 'Gurugram', 'panipat': 'Panipat Rural'}
pairs, unmatched = [], []
used = set()
for f in pdfs:
    n = norm(f)
    m = ALIASES.get(n) or sheet_by_norm.get(n)
    if not m:
        cand = difflib.get_close_matches(n, [k for k in sheet_by_norm if sheet_by_norm[k] not in used], n=1, cutoff=0.75)
        m = sheet_by_norm[cand[0]] if cand else None
    if m:
        pairs.append((f, m)); used.add(m)
    else:
        unmatched.append(f)
print('matched:', len(pairs), 'unmatched pdfs:', unmatched)
print('unused sheets:', [s for s in sheets if s not in used])

def get_pages_text(doc, cache_key=None):
    pages = [p.get_text() for p in doc]
    if sum(len(t) for t in pages) > 200 * len(pages):
        return pages, False
    cache = OUT + 'ocr_cache/' + (cache_key or 'x') + '.json'
    if cache_key and os.path.exists(cache):
        return json.load(open(cache)), True
    raise RuntimeError('needs OCR, no cache: ' + str(cache_key))

TOK = re.compile(r'^\d+[A-Z]?$')

def consume(toks, W):
    i, expect, rows = 0, 1, []
    while i < len(toks):
        t = toks[i]
        if isinstance(t, str):
            pre = str(expect)
            if t.startswith(pre) and re.fullmatch(r'\d+[A-Z]', t[len(pre):]):
                serial, ps, j = expect, t[len(pre):], i + 1
            else:
                break
        elif t == expect:
            if i + 1 >= len(toks):
                break
            ps = toks[i+1]
            if isinstance(ps, str) and not re.fullmatch(r'\d+[A-Z]', ps):
                break
            serial, j = expect, i + 2
        else:
            break
        body = toks[j:j+W-2]
        if len(body) < W - 2 or not all(isinstance(x, int) for x in body):
            break
        rows.append([serial, ps] + body)
        i = j + W - 2
        expect += 1
    return rows, toks[i:]

def parse_form20(path):
    if os.path.getsize(path) < 5000:
        return {'error': 'file is a broken download (404 page), re-download from CEO Haryana'}
    doc = fitz.open(path)
    pages, ocred = get_pages_text(doc, os.path.basename(path).replace('.pdf', ''))
    lines0 = pages[0].split('\n')
    ac_name = ''
    for l in lines0:
        if 'Name of Assembly' in l:
            ac_name = l.split('...')[-1].strip()
    # candidate names: between last header 'Votes' line and first numeric line
    i_tv = [i for i, l in enumerate(lines0) if l.strip() == 'Votes']
    start = i_tv[-1] + 1 if i_tv else 0
    names_raw = []
    for l in lines0[start:]:
        if re.fullmatch(r'\d+(\s+\d+)*', l.strip()):
            break
        names_raw.append(l.strip())
    header_text = ' '.join(w for w in names_raw if w)

    flat = []
    for pt in pages:
        for l in pt.split('\n'):
            toks = l.split()
            if toks and all(TOK.fullmatch(t) for t in toks):
                flat.extend(int(t) if t.isdigit() else t for t in toks)
    best = None
    for w in range(8, 45):
        rows, rest = consume(flat, w)
        if best is None or len(rows) > len(best[1]):
            best = (w, rows, rest)
    W, rows, leftover = best
    leftover = [x for x in leftover if isinstance(x, int)]
    ncand = W - 7
    if ncand < 2 or len(rows) < 20:
        return {'error': f'parse failed (width={W}, rows={len(rows)})'}
    internal_bad = 0
    for r in rows:
        if sum(r[2:2+ncand]) != r[2+ncand] or r[2+ncand]+r[3+ncand]+r[4+ncand] != r[5+ncand]:
            internal_bad += 1
    # leftover: EVM totals (ncand+5), postal (ncand+5), grand (ncand+5)
    L = ncand + 5
    evm = leftover[0:L] if len(leftover) >= L else None
    postal = leftover[L:2*L] if len(leftover) >= 2*L else None
    grand = leftover[2*L:3*L] if len(leftover) >= 3*L else None
    booth_sums = [sum(r[2+i] for r in rows) for i in range(ncand)]
    evm_match = (evm is not None and booth_sums == evm[:ncand]
                 and sum(r[2+ncand] for r in rows) == evm[ncand])
    return {'ac_name': ac_name, 'ncand': ncand, 'rows': rows, 'header': header_text,
            'ocred': ocred, 'internal_bad': internal_bad, 'evm_match': evm_match,
            'evm': evm, 'postal': postal, 'grand': grand,
            'leftover_len': len(leftover), 'expected_leftover': 3*L}

report = []
master_rows = []   # long format: ac, village, candidate_idx, candidate_name?, votes
village_summary = []
cand_meta = {}

os.makedirs(OUT + 'parse_cache', exist_ok=True)
def parse_cached(f):
    cf = OUT + 'parse_cache/' + f.replace('.pdf', '') + '.json'
    if os.path.exists(cf):
        return json.load(open(cf))
    res = parse_form20(PDF_DIR + f)
    json.dump(res, open(cf, 'w'))
    return res

def evm_check(res):
    if res['evm_match']:
        return True
    if not res['evm']:
        return False
    nc = res['ncand']
    rows = res['rows']
    booth_sums = [sum(r[2+i] for r in rows) for i in range(nc)]
    valid_sum = sum(r[2+nc] for r in rows)
    expected = ''.join(str(x) for x in booth_sums + [valid_sum])
    actual = ''.join(str(x) for x in res['evm'][:nc+1])
    return actual.startswith(expected)  # tolerates concatenated tokens in totals row

for f, sheet in pairs:
    res = parse_cached(f)
    ac = sheet
    if 'error' in res:
        report.append({'ac': ac, 'pdf': f, 'status': 'ERROR', 'note': res['error']})
        continue
    res['evm_match'] = evm_check(res)
    # booth->village
    # user-supplied booth-range fixes (2026-08-02)
    RANGE_FIXES = {
        'Badkhal': [(189, 228, 'Lakadpur'), (229, 256, 'Anangpur'), (257, 283, 'Fatehpur Chandila')],
        'Faridabad': [(154, 165, 'Daulatabad'), (166, 187, 'Ajraunda'), (188, 244, 'Sihi'), (245, 249, 'Ballabgarh')],
        'Kaithal': [(1,4,'Niwach'),(5,6,'Balwanti'),(7,8,'Jaswanti'),(9,18,'Kyodak'),(19,21,'Dayora'),
            (22,22,'Ujhana'),(23,23,'Jagdish Pura'),(24,25,'Kultaran'),(26,28,'Khurana'),
            (29,31,'Patti Afghan (Urban)'),(32,34,'Sirta'),(35,38,'Manas'),(39,41,'Ladana Baba'),
            (42,44,'Budhakhera'),(45,45,'Sangatpura'),(46,46,'Nand Singh Wala'),(47,49,'Sanghan'),
            (50,51,'Malkhedi'),(52,55,'Padla'),(56,56,'Chakk Padla'),(57,57,'Diluwali'),(58,61,'Guhna'),
            (62,66,'Sajuma'),(67,68,'Dundrehedi'),(69,70,'Diwal'),(71,72,'Chhot'),(73,73,'Bhanpura'),
            (74,74,'Gadi Padla'),(75,75,'Madho Majri'),(76,77,'Patti Khot / Gadli'),(78,78,'Phansawala'),
            (79,80,'Kutubpur'),(81,81,'Patti Dogar / Shila Khera'),(82,90,'Arjun Nagar / Sirta Road'),
            (91,94,'Shakti Nagar'),(95,96,'Balaji Colony / Bank Colony / Rajni Colony'),
            (97,97,'Devigarh / Shiv Nagar'),(98,98,'Friends Colony / HUDA Sector 18'),(99,100,'Balraj Nagar'),
            (101,104,'Subhash Nagar / Friends Colony / Janakpuri'),(105,107,'Mayapuri / Sugar Mill Colony'),
            (108,110,'Nankpuri Colony / D.P.V. Colony'),(111,113,'HUDA Sector 19 / Sector 20 / Rishi Nagar'),
            (114,116,'HUDA Sector 21 / Rajouri Garden / Moti Bagh'),(117,118,'HUDA Sector 19 / Officer Colony'),
            (119,120,'HUDA Sector 20'),(121,123,'Siwan Gate / Dogran Gate'),(124,127,'Pratap Gate / Mata Gate'),
            (128,129,'Mahadev Colony / Rajiv Colony'),(130,134,'West Bihar Colony / Gupta Colony / Subhash Nagar'),
            (135,137,'Agrasen Puram / RK Puram / Employees Colony'),(138,139,'Chiranjeev Colony / Seth Colony'),
            (140,141,'Khushhal Majri / Chichdan Mohalla'),(142,143,'Jain Mohalla / Joshian Mohalla'),
            (144,146,'Main Bazar / Shastri Market / Prem Gali'),(147,148,'Sivka Market / Railway Gate'),
            (149,151,'State Bank Colony / GTB Colony / Govind Nagar'),(152,152,'Canal Colony / MITC Colony'),
            (153,155,'Adarsh Nagar / Professor Colony'),(156,156,'Model Town / PWD Colony'),
            (157,162,'Amargadh Gamri / Kamal Colony / Krishna Nagar'),
            (163,166,'Patel Nagar / Sarsoda Colony / Om Shanti Nagar'),
            (167,169,'Model Town Jind Road / Sora Kothi'),(170,171,'Bank Colony / Ram Nagar'),
            (172,175,'HUDA Housing Board / Chanda Road'),(176,180,'Saini Colony / Gabi Sahib Colony'),
            (181,183,'Sripunj Mohalla / Khurana Mohalla'),(184,188,'Pratap Gate / Ambkeshwar Colony'),
            (189,190,'Shiv Nagar / Azad Nagar'),(191,191,'Shergad'),(192,193,'Dayodkhedi'),
            (194,194,'Bhaini Majra'),(195,198,'Gyong'),(199,199,'Sapan Khedi'),(200,203,'Munddi'),
            (204,206,'Naina'),(207,209,'Kathwad'),(210,212,'Dhaus'),(213,215,'Khanoda')],
    }
    SHEET_SKIP = {'Kaithal'}  # sheet contains wrong (Pundri ward) data; use range fixes only
    b2v = {}
    for a, b, vname in RANGE_FIXES.get(sheet, []):
        for k in range(a, b + 1):
            b2v[k] = vname
    for r in (() if sheet in SHEET_SKIP else wb_map[sheet].iter_rows(min_row=2, values_only=True)):
        if r[0] is not None and r[1] is not None:
            key = r[0]
            try:
                key = int(key)
            except (ValueError, TypeError):
                key = str(key).strip().upper()
            if key not in b2v:                 # keep the FIRST village listed
                b2v[key] = str(r[1]).strip()

    def lookup(ps):
        if ps in b2v:
            return b2v[ps]
        if isinstance(ps, str):  # auxiliary booth like '160A' -> parent 160
            m = re.match(r'\d+', ps)
            if m and int(m.group()) in b2v:
                return b2v[int(m.group())]
        return None

    rows = res['rows']; nc = res['ncand']
    unmapped = sorted({str(r[1]) for r in rows if lookup(r[1]) is None})
    mapped_no_booth = sorted(str(k) for k in (set(b2v) - {r[1] for r in rows}) if isinstance(k, int))
    agg = defaultdict(lambda: [0]*(nc+3))  # votes..., valid, nota, booths
    for r in rows:
        v = lookup(r[1]) or 'UNMAPPED'
        a = agg[v]
        for i in range(nc):
            a[i] += r[2+i]
        a[nc] += r[2+nc]; a[nc+1] += r[4+nc]; a[nc+2] += 1
    cand_meta[ac] = {'header': res['header'], 'ncand': nc,
                     'evm': res['evm'], 'postal': res['postal'], 'grand': res['grand']}
    for v, a in agg.items():
        village_summary.append({'ac': ac, 'village': v, 'booths': a[nc+2],
                                'valid': a[nc], 'nota': a[nc+1]})
        for i in range(nc):
            master_rows.append((ac, v, i, a[i]))
    report.append({'ac': ac, 'pdf': f, 'status': 'OK', 'ocred': res['ocred'],
                   'booths': len(rows), 'ncand': nc,
                   'internal_bad': res['internal_bad'],
                   'evm_match': res['evm_match'],
                   'unmapped_booths': len(unmapped),
                   'unmapped_list': unmapped[:20],
                   'excel_booths_not_in_pdf': len(mapped_no_booth),
                   'villages': len(agg)})

ok = [r for r in report if r['status'] == 'OK']
perfect = [r for r in ok if r['internal_bad'] == 0 and r['evm_match'] and r['unmapped_booths'] == 0]
print('\n=== SUMMARY: %d/%d parsed | %d fully clean ===' % (len(ok), len(pairs), len(perfect)))
for r in report:
    flag = '' if (r['status'] == 'OK' and r['internal_bad'] == 0 and r['evm_match'] and r['unmapped_booths'] == 0) else '  <<< CHECK'
    if r['status'] == 'OK':
        print('%-24s booths=%4d cand=%2d villages=%3d badrows=%d evm_match=%s unmapped=%d%s' % (
            r['ac'], r['booths'], r['ncand'], r['villages'], r['internal_bad'], r['evm_match'], r['unmapped_booths'], flag))
    else:
        print('%-24s ERROR: %s%s' % (r['ac'], r['note'], flag))

with open(OUT + 'batch_report.json', 'w') as fo:
    json.dump(report, fo, indent=1)
with open(OUT + 'cand_meta.json', 'w') as fo:
    json.dump(cand_meta, fo, indent=1)
with open(OUT + 'village_master.csv', 'w', newline='') as fo:
    w = csv.writer(fo)
    w.writerow(['assembly', 'village', 'candidate_index', 'votes'])
    w.writerows(master_rows)
with open(OUT + 'village_summary.csv', 'w', newline='') as fo:
    w = csv.writer(fo)
    w.writerow(['assembly', 'village', 'booths', 'valid_votes', 'nota'])
    for v in village_summary:
        w.writerow([v['ac'], v['village'], v['booths'], v['valid'], v['nota']])
print('\nmaster rows:', len(master_rows), '| village rows:', len(village_summary))
