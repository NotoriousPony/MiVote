"""Validation report workbook for all 90 ACs."""
import json, csv, sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/sessions/inspiring-charming-cerf/mnt/outputs/'
report = json.load(open(OUT + 'batch_report.json'))
con = sqlite3.connect('/tmp/haryana_village_results_2024.db')

wb = openpyxl.Workbook()
F = 'Arial'
hdr_fill = PatternFill('solid', fgColor='1F4E79')
warn_fill = PatternFill('solid', fgColor='FCE4EC')
ok_fill = PatternFill('solid', fgColor='E8F5E9')
thin = Border(*[Side(style='thin', color='BFBFBF')]*4)

def header(ws, cols):
    ws.append(cols)
    for c in range(1, len(cols)+1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(name=F, bold=True, color='FFFFFF', size=10)
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'

ws = wb.active
ws.title = 'AC Status'
header(ws, ['#', 'Assembly', 'Booths (Form 20)', 'Candidates', 'Villages', 'Rows failing checksum',
            'Totals match Form 20', 'Unmapped booths', 'Status / action needed'])
for i, e in enumerate(sorted(report, key=lambda x: x['ac']), start=1):
    if e['status'] != 'OK':
        ws.append([i, e['ac'], None, None, None, None, None, None, e['note']])
        bad = True
    else:
        action = 'OK'
        bad = False
        if e['unmapped_booths']:
            action = f"Add booths to Excel mapping ({e['unmapped_booths']} missing: {', '.join(e['unmapped_list'][:6])}...)"
            bad = True
        ws.append([i, e['ac'], e['booths'], e['ncand'], e['villages'], e['internal_bad'],
                   'YES' if e['evm_match'] else 'NO', e['unmapped_booths'], action])
    rn = ws.max_row
    for c in range(1, 10):
        cell = ws.cell(row=rn, column=c)
        cell.font = Font(name=F, size=10)
        cell.border = thin
        cell.fill = warn_fill if bad else ok_fill
ws.column_dimensions['B'].width = 20
for col in 'CDEFGH':
    ws.column_dimensions[col].width = 12
ws.column_dimensions['I'].width = 55

# candidate proofread sheet
ws2 = wb.create_sheet('Candidates (proofread names)')
header(ws2, ['Assembly', 'Rank', 'Candidate name (from PDF layout)', 'Party (fill in)', 'Total EVM votes', 'Top 4?'])
rows = con.execute('''SELECT ac_name, rank, name, total_votes_evm, is_top4
                      FROM candidate ORDER BY ac_name, rank''').fetchall()
for ac, rank, name, votes, top4 in rows:
    ws2.append([ac, rank, name, None, votes, 'YES' if top4 else ''])
    rn = ws2.max_row
    for c in range(1, 7):
        cell = ws2.cell(row=rn, column=c)
        cell.font = Font(name=F, size=10)
        cell.border = thin
        if c == 5:
            cell.number_format = '#,##0'
ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['C'].width = 32
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 14

ws3 = wb.create_sheet('Notes')
notes = [
    'HARYANA 2024 ASSEMBLY - VILLAGE-WISE RESULTS: DATA VALIDATION REPORT',
    '',
    'Source: Form 20 Final Result Sheets (CEO Haryana), 90 ACs; booth-to-village mapping from Haryana_Assembly_Normalized.xlsx (user-provided).',
    'Validation: every booth row checksummed (candidate votes = valid total; valid+rejected+NOTA = total), and per-candidate booth sums matched against the Form 20 printed EVM totals row. 89/89 parsed ACs match exactly.',
    '',
    'RESOLVED: Badkhal, Faridabad and Kaithal booth-village mappings supplied by user (2 Aug 2026) and applied.',
    'Shahbad (SC): Form 20 not released by government; database carries AC-level totals from ECI (Ram Karan INC 61,050 def. Subhash Chand BJP 54,609) with status AC_TOTAL_ONLY. App must show the overall result with footer: "Government is yet to release Form 20 (booth-wise) data for this constituency."',
    '',
    'Party labels applied from user-supplied alliance list (342/343 matched; 1 withdrawal). Ratia BSP candidate (Chhindwara Pal) withdrew - confirmed by user. Jind JJP Dharampal Prajapat appears on ballot as Dharam Pal Tanwar - confirmed same person by user.',
    'REMAINING: proofread candidate names in the Candidates sheet for minor spelling issues.',
    '',
    'Postal ballots are counted at AC level and are NOT included in village-wise numbers (they cannot be attributed to villages). Show them separately in the app.',
    'Auxiliary booths (e.g. 160A) are credited to the same village as their parent booth.',
]
for i, n in enumerate(notes, start=1):
    ws3.cell(row=i, column=1, value=n).font = Font(name=F, size=10, bold=(i == 1))
ws3.column_dimensions['A'].width = 120

wb.save(OUT + 'Haryana_2024_Validation_Report.xlsx')
print('saved report;', ws.max_row-1, 'ACs;', ws2.max_row-1, 'candidate rows')
